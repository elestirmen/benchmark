#!/usr/bin/env python
"""Build and validate the benchmark workbook with the approved openpyxl fallback.

The canonical benchmark data remains JSON/JSONL/CSV.  This reporter creates a
human-readable Excel view without changing any scientific measurements.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import math
import os
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo


LOG = logging.getLogger("benchmark_excel")
MAX_EXCEL_ROWS = 1_048_576
MAX_CELL_TEXT = 32_700
STATE_SHEET = "_ExcelState"
STATE_SCHEMA_VERSION = 1
LATEST_POINTER = "excel_latest.json"
REQUIRED_SHEETS = [
    "Özet",
    "Model Özeti",
    "Model Kataloğu",
    "Ham Sonuçlar",
    "Sorgu Manifesti",
    "Yapılandırma",
    "Hatalar",
]
LIGHTWEIGHT_SHEETS = [
    "Özet",
    "Model Özeti",
    "Model Kataloğu",
    "Yapılandırma",
    "Hatalar",
]

DARK = "374151"
DARKER = "1F2937"
MID = "6B7280"
LIGHT = "E5E7EB"
LIGHTER = "F9FAFB"
ACCENT = "0F766E"
GOOD = "DCFCE7"
WARN = "FEF3C7"
BAD = "FEE2E2"
WHITE = "FFFFFF"
TEXT = "111827"

THIN_GRAY = Side(style="thin", color="D1D5DB")
SUBTLE_BORDER = Border(bottom=Side(style="thin", color=LIGHT))

SUMMARY_COLUMNS = [
    "direction",
    "query_variant",
    "search_mode",
    "model_id",
    "model_label",
    "model_epoch",
    "model_relative_path",
    "model_sha256",
    "total_queries",
    "ok_queries",
    "rejected_queries",
    "error_queries",
    "coverage",
    "success_25m_queries",
    "success_25m",
    "success_25m_ci95_low",
    "success_25m_ci95_high",
    "success_25m_failure_rate",
    "auc_25m",
    "mean_error_under_25m",
    "median_error_under_25m",
    "mean_error_m",
    "median_error_m",
    "median_error_ci95_low",
    "median_error_ci95_high",
    "p90_error_m",
    "p95_error_m",
    "success_5m",
    "success_10m",
    "success_10m_ci95_low",
    "success_10m_ci95_high",
    "success_50m",
    "mean_top1_score",
    "mean_search_seconds",
    "total_search_seconds",
]

# SUMMARY_COLUMNS remains the canonical complete schema for JSON/CSV compatibility.
# This compact projection is the human-facing Excel view.
MODEL_SUMMARY_COLUMNS = [
    "model_sequence",
    "direction",
    "query_variant",
    "search_mode",
    "model_label",
    "model_id",
    "model_epoch",
    "total_queries",
    "ok_queries",
    "coverage",
    "success_25m",
    "auc_25m",
    "median_error_under_25m",
    "p90_error_m",
    "mean_search_seconds",
]

MODEL_SUMMARY_HEADERS = {
    "model_sequence": "Sıra No",
    "direction": "Yön",
    "query_variant": "Senaryo",
    "search_mode": "Arama",
    "model_label": "Model",
    "model_id": "Model ID",
    "model_epoch": "Epoch",
    "total_queries": "N",
    "ok_queries": "Başarılı N",
    "coverage": "Kapsam",
    "success_25m": "Başarı ≤25 m",
    "auc_25m": "AUC@25 m",
    "median_error_under_25m": "Medyan hata ≤25 m (m)",
    "p90_error_m": "P90 hata (m)",
    "mean_search_seconds": "Ort. arama (sn)",
}

MODEL_CATALOG_COLUMNS = [
    "model_id",
    "model_label",
    "model_epoch",
    "model_relative_path",
    "model_sha256",
]

MODEL_CATALOG_HEADERS = {
    "model_id": "Model ID (kanonik)",
    "model_label": "Model",
    "model_epoch": "Epoch",
    "model_relative_path": "Göreli yol",
    "model_sha256": "SHA256",
}

QUERY_COLUMNS = [
    "direction",
    "query_variant",
    "query_id",
    "block_id",
    "center_easting_m",
    "center_northing_m",
    "source_row",
    "source_col",
    "query_std",
    "query_entropy",
    "dark_fraction",
    "raw_tile_file",
]

SUMMARY_FORMATS = {
    "coverage": "0.0%",
    "success_25m": "0.0%",
    "success_25m_ci95_low": "0.0%",
    "success_25m_ci95_high": "0.0%",
    "success_25m_failure_rate": "0.0%",
    "auc_25m": "0.0%",
    "mean_error_under_25m": "0.00",
    "median_error_under_25m": "0.00",
    "mean_error_m": "0.00",
    "median_error_m": "0.00",
    "median_error_ci95_low": "0.00",
    "median_error_ci95_high": "0.00",
    "p90_error_m": "0.00",
    "p95_error_m": "0.00",
    "success_5m": "0.0%",
    "success_10m": "0.0%",
    "success_10m_ci95_low": "0.0%",
    "success_10m_ci95_high": "0.0%",
    "success_50m": "0.0%",
    "mean_top1_score": "0.0000",
    "mean_search_seconds": "0.000",
    "total_search_seconds": "0.0",
}

RESULT_FORMATS = {
    "query_center_easting_m": "0.000",
    "query_center_northing_m": "0.000",
    "expected_center_easting_m": "0.000",
    "expected_center_northing_m": "0.000",
    "predicted_center_easting_m": "0.000",
    "predicted_center_northing_m": "0.000",
    "error_m": "0.000",
    "error_px": "0.000",
    "top1_score": "0.0000",
    "top2_score": "0.0000",
    "peak_margin": "0.0000",
    "psr": "0.00",
    "query_std": "0.00",
    "query_entropy": "0.00",
    "query_dark_fraction": "0.0%",
    "success_25m": "0",
    "search_seconds": "0.000",
    "query_inference_seconds": "0.000",
    "map_build_seconds": "0.0",
}

QUERY_FORMATS = {
    "center_easting_m": "0.000",
    "center_northing_m": "0.000",
    "query_std": "0.00",
    "query_entropy": "0.00",
    "dark_fraction": "0.0%",
}


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Benchmark sonuçlarını XLSX raporuna dönüştürür.")
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument(
        "--incremental",
        action="store_true",
        help="Var olan rapora yalnızca yeni ham sonuç satırlarını ekler.",
    )
    parser.add_argument(
        "--lightweight",
        action="store_true",
        help="Model sonunda yalnız özet, sıralama, yapılandırma ve hataları yazar.",
    )
    parser.add_argument(
        "--validation-mode",
        choices=("checkpoint", "deep"),
        default="deep",
        help="Ara model raporunda hızlı, final raporda derin doğrulama.",
    )
    return parser.parse_args(argv)


class IncrementalUpdateUnavailable(RuntimeError):
    """Raised when a safe append cannot be proven and a full rebuild is required."""


def read_json(path: Path, fallback: Any) -> Any:
    if not path.is_file():
        return fallback
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            line = line.strip()
            if not line:
                continue
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"{path}:{line_number} bir JSON nesnesi değil.")
            rows.append(value)
    return rows


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    if not path.is_file():
        return digest.hexdigest()
    with path.open("rb") as handle:
        while True:
            block = handle.read(8 * 1024 * 1024)
            if not block:
                break
            digest.update(block)
    return digest.hexdigest()


def source_identity(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {"exists": False, "size_bytes": 0, "sha256": sha256_file(path)}
    stat = path.stat()
    return {
        "exists": True,
        "size_bytes": stat.st_size,
        "sha256": sha256_file(path),
    }


def query_manifest_identity(run_dir: Path) -> str:
    digest = hashlib.sha256()
    paths = sorted(run_dir.rglob("query_manifest.json")) + sorted(
        run_dir.rglob("query_variant_manifest.json")
    )
    for path in paths:
        digest.update(path.relative_to(run_dir).as_posix().encode("utf-8"))
        digest.update(b"\0")
        digest.update(path.read_bytes())
        digest.update(b"\0")
    return digest.hexdigest()


def build_excel_state(
    run_dir: Path,
    *,
    result_rows: int,
    result_columns: Sequence[str],
    results_identity: dict[str, Any] | None = None,
) -> dict[str, Any]:
    results_path = run_dir / "results.jsonl"
    return {
        "schema_version": STATE_SCHEMA_VERSION,
        "updated_at_utc": utc_now_iso(),
        "result_rows": int(result_rows),
        "result_columns": list(result_columns),
        "results": results_identity or source_identity(results_path),
        "summary": source_identity(run_dir / "summary.json"),
        "config": source_identity(run_dir / "run_config.json"),
        "errors": source_identity(run_dir / "model_errors.jsonl"),
        "query_manifests_sha256": query_manifest_identity(run_dir),
    }


def write_excel_state(wb: Workbook, state: dict[str, Any]) -> None:
    if STATE_SHEET in wb.sheetnames:
        wb.remove(wb[STATE_SHEET])
    ws = wb.create_sheet(STATE_SHEET)
    ws["A1"] = "benchmark_excel_state_json"
    ws["A2"] = json.dumps(state, ensure_ascii=False, sort_keys=True)
    ws.sheet_state = "veryHidden"


def read_excel_state(wb: Workbook) -> dict[str, Any] | None:
    if STATE_SHEET not in wb.sheetnames:
        return None
    value = wb[STATE_SHEET]["A2"].value
    if not isinstance(value, str):
        return None
    try:
        state = json.loads(value)
    except json.JSONDecodeError:
        return None
    return state if isinstance(state, dict) else None


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    text = str(value)
    if len(text) > MAX_CELL_TEXT:
        text = text[: MAX_CELL_TEXT - 25] + " … [kısaltıldı]"
    return text


def excel_values_equal(actual: Any, expected: Any) -> bool:
    # OOXML/openpyxl round-trips an empty JSON string cell as an empty cell.
    if actual in (None, "") and expected in (None, ""):
        return True
    numeric_types = (int, float)
    if (
        isinstance(actual, numeric_types)
        and not isinstance(actual, bool)
        and isinstance(expected, numeric_types)
        and not isinstance(expected, bool)
    ):
        if isinstance(actual, int) and isinstance(expected, int):
            return actual == expected
        # Excel stores at most about 15 significant decimal digits.  JSONL stays canonical.
        return math.isclose(float(actual), float(expected), rel_tol=5e-15, abs_tol=5e-15)
    return actual == expected


def ordered_columns(rows: Sequence[dict[str, Any]], fallback: Sequence[str]) -> list[str]:
    seen: set[str] = set()
    columns: list[str] = []
    for name in fallback:
        if name not in seen:
            seen.add(name)
            columns.append(name)
    for row in rows:
        for name in row:
            if name not in seen:
                seen.add(name)
                columns.append(name)
    return columns


def enrich_summary_model_identity(
    run_dir: Path, rows: Sequence[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Add human-readable epoch identity without changing canonical model IDs."""
    catalog = read_json(run_dir / "model_catalog.json", {})
    models = catalog.get("models", []) if isinstance(catalog, dict) else []
    sequence_by_model_id = {
        str(item["model_id"]): index + 1
        for index, item in enumerate(models)
        if isinstance(item, dict) and item.get("model_id")
    }
    by_model_id = {
        str(item["model_id"]): item
        for item in models
        if isinstance(item, dict) and item.get("model_id")
    }
    enriched: list[dict[str, Any]] = []
    for source in rows:
        row = dict(source)
        model_id = str(row.get("model_id", ""))
        if model_id == "RAW_BASELINE":
            row.update(
                {
                    "model_label": "RAW_BASELINE",
                    "model_sequence": 0,
                    "model_epoch": None,
                    "model_relative_path": None,
                    "model_sha256": None,
                }
            )
            enriched.append(row)
            continue
        identity = by_model_id.get(model_id, {})
        if not identity:
            row.update(
                {
                    "model_label": model_id,
                    "model_sequence": sequence_by_model_id.get(model_id),
                    "model_epoch": None,
                    "model_relative_path": None,
                    "model_sha256": None,
                }
            )
            enriched.append(row)
            continue
        relative_path = str(identity.get("relative_path") or "")
        checkpoint = identity.get("checkpoint_number")
        parent_name = Path(relative_path).parent.name if relative_path else ""
        rank_match = re.match(r"^(\d{2})_", parent_name)
        lineage_label = (
            f"Lineage {rank_match.group(1)}"
            if rank_match
            else (parent_name or "Model")
        )
        epoch_label = (
            f"epoch_{int(checkpoint):05d}"
            if checkpoint is not None
            else Path(relative_path).stem
        )
        row.update(
            {
                "model_label": f"{lineage_label} | {epoch_label}",
                "model_sequence": sequence_by_model_id.get(model_id),
                "model_epoch": int(checkpoint) if checkpoint is not None else None,
                "model_relative_path": relative_path or None,
                "model_sha256": identity.get("sha256"),
            }
        )
        enriched.append(row)
    return enriched


def display_direction(value: Any) -> Any:
    if value is None:
        return None
    text = str(value)
    lowered = text.lower()
    gmaps_index = lowered.find("gmaps")
    bing_index = lowered.find("bingmap")
    if gmaps_index >= 0 and bing_index >= 0:
        return "Google → Bing" if gmaps_index < bing_index else "Bing → Google"
    return text.replace("__TO__", " → ")


def compact_direction_label(value: Any) -> str:
    text = str(value or "").lower()
    gmaps_index = text.find("gmaps") if "gmaps" in text else text.find("google")
    bing_index = text.find("bingmap") if "bingmap" in text else text.find("bing")
    if gmaps_index >= 0 and bing_index >= 0:
        return "G→B" if gmaps_index < bing_index else "B→G"
    return "Yön"


def display_query_variant(value: Any) -> Any:
    if value is None:
        return None
    return {"clean": "Temiz", "hard_v1": "Hard"}.get(str(value), value)


def display_search_mode(value: Any) -> Any:
    if value is None:
        return None
    text = str(value)
    if text == "global":
        return "Global"
    match = re.fullmatch(r"roi[_-]?(\d+)(?:m)?", text, flags=re.IGNORECASE)
    if match:
        return f"ROI {int(match.group(1))} m"
    return text


def prepare_model_summary_rows(rows: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    prepared: list[dict[str, Any]] = []
    for source in rows:
        row = dict(source)
        row["direction"] = display_direction(row.get("direction"))
        row["query_variant"] = display_query_variant(row.get("query_variant"))
        row["search_mode"] = display_search_mode(row.get("search_mode"))
        prepared.append(row)
    prepared.sort(
        key=lambda row: (
            row.get("model_sequence") is None,
            row.get("model_sequence") if row.get("model_sequence") is not None else float("inf"),
        )
    )
    return prepared


def compact_model_label(row: dict[str, Any]) -> str:
    model_label = str(row.get("model_label") or row.get("model_id") or "Model")
    if model_label == "RAW_BASELINE":
        return "RAW"
    match = re.search(r"Lineage\s+(\d+)", model_label)
    epoch = row.get("model_epoch")
    if match and epoch is not None:
        return f"L{int(match.group(1)):02d} | E{int(epoch):03d}"
    if epoch is not None:
        return f"{model_label[:14]} | E{int(epoch):03d}"
    return model_label[:24]


def build_model_catalog_rows(
    run_dir: Path, summary_rows: Sequence[dict[str, Any]]
) -> list[dict[str, Any]]:
    catalog = read_json(run_dir / "model_catalog.json", {})
    models = catalog.get("models", []) if isinstance(catalog, dict) else []
    model_ids: list[str] = []
    for item in models:
        if isinstance(item, dict) and item.get("model_id"):
            model_ids.append(str(item["model_id"]))
    for row in summary_rows:
        if row.get("model_id"):
            model_ids.append(str(row["model_id"]))
    unique_ids = list(dict.fromkeys(model_ids))
    enriched = enrich_summary_model_identity(
        run_dir, [{"model_id": model_id} for model_id in unique_ids]
    )
    return [
        {column: row.get(column) for column in MODEL_CATALOG_COLUMNS}
        for row in enriched
    ]


def find_query_rows(run_dir: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    manifest_paths = sorted(run_dir.rglob("query_manifest.json")) + sorted(
        run_dir.rglob("query_variant_manifest.json")
    )
    for manifest_path in manifest_paths:
        manifest = read_json(manifest_path, {})
        try:
            direction = manifest_path.relative_to(run_dir).parts[0]
        except IndexError:
            direction = "unknown"
        query_variant = str(manifest.get("query_variant", "clean"))
        for query in manifest.get("queries", []):
            if isinstance(query, dict):
                row = {"direction": direction, "query_variant": query_variant, **query}
                if isinstance(row.get("augmentation"), dict):
                    row["augmentation"] = json.dumps(
                        row["augmentation"], ensure_ascii=False, sort_keys=True
                    )
                rows.append(row)
    return rows


def safe_table_name(prefix: str, index: int) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", prefix)
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return f"{cleaned}_{index:02d}"


def configure_sheet(
    ws, *, freeze: str = "A2", fit_width: int = 1, orientation: str = "portrait"
) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = orientation
    ws.print_title_rows = "1:1"
    ws.sheet_view.zoomScale = 90


def style_header(row: Iterable[Any]) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=MID))


def style_data_rows(
    ws,
    columns: Sequence[str],
    *,
    min_row: int,
    max_row: int,
    number_formats: dict[str, str] | None = None,
) -> None:
    if max_row < min_row:
        return
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        for cell in row:
            cell.font = Font(name="Aptos", size=9, color=TEXT)
            cell.alignment = Alignment(vertical="center")
            cell.border = SUBTLE_BORDER
    if number_formats:
        for index, column in enumerate(columns, start=1):
            number_format = number_formats.get(column)
            if number_format:
                for row in range(min_row, max_row + 1):
                    cell = ws.cell(row=row, column=index)
                    cell.number_format = number_format
                    cell.alignment = Alignment(horizontal="right", vertical="center")


def write_table(
    ws,
    rows: Sequence[dict[str, Any]],
    columns: Sequence[str],
    *,
    table_name: str,
    number_formats: dict[str, str] | None = None,
    headers: Sequence[str] | None = None,
) -> None:
    if headers is not None and len(headers) != len(columns):
        raise ValueError("Excel başlıkları sütun sayısıyla aynı olmalıdır.")
    ws.append(list(headers or columns))
    for row in rows:
        ws.append([excel_value(row.get(column)) for column in columns])
    # Excel desktop rejects a table whose reference contains only its header.
    # Keep a single genuinely blank data row for empty datasets.
    if not rows:
        ws.append([None] * len(columns))
    style_header(ws[1])
    ws.row_dimensions[1].height = 32
    style_data_rows(
        ws,
        columns,
        min_row=2,
        max_row=max(2, ws.max_row),
        number_formats=number_formats,
    )
    end_column = get_column_letter(len(columns))
    table = Table(displayName=table_name, ref=f"A1:{end_column}{max(1, ws.max_row)}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    column_count = len(columns)
    configure_sheet(
        ws,
        fit_width=3 if column_count > 30 else 2 if column_count > 15 else 1,
        orientation="landscape" if column_count > 6 else "portrait",
    )


def set_widths(
    ws,
    columns: Sequence[str],
    *,
    overrides: dict[str, float] | None = None,
    sample_rows: int = 250,
) -> None:
    overrides = overrides or {}
    for column_index, name in enumerate(columns, start=1):
        if name in overrides:
            width = overrides[name]
        else:
            values = [str(name)]
            for row_index in range(2, min(ws.max_row, sample_rows + 1) + 1):
                value = ws.cell(row=row_index, column=column_index).value
                if value is not None:
                    values.append(str(value))
            width = min(24.0, max(10.0, max(len(value) for value in values) + 2.0))
        ws.column_dimensions[get_column_letter(column_index)].width = width


def add_error_conditional_formatting(
    ws, columns: Sequence[str], *, min_row: int = 2, max_row: int | None = None
) -> None:
    max_row = ws.max_row if max_row is None else max_row
    if max_row < min_row:
        return
    if "error_m" in columns:
        index = columns.index("error_m") + 1
        letter = get_column_letter(index)
        ws.conditional_formatting.add(
            f"{letter}{min_row}:{letter}{max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="DCFCE7",
                mid_type="percentile",
                mid_value=50,
                mid_color="FEF3C7",
                end_type="max",
                end_color="FECACA",
            ),
        )


def write_dashboard(
    wb: Workbook,
    config: dict[str, Any],
    summary_rows: Sequence[dict[str, Any]],
    result_chunks: Sequence[tuple[str, int, Sequence[str]]],
    *,
    lightweight: bool = False,
) -> None:
    ws = wb["Özet"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 90
    ws.merge_cells("A1:H2")
    ws["A1"] = "Jeoreferanslı Ortak-Temsil Benchmark Sonuçları"
    ws["A1"].fill = PatternFill("solid", fgColor=DARKER)
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    labels = ["Koşu", "Toplam kayıt", "Model/yön kombinasyonu", "Başarılı kayıt", "Hata/reddetme"]
    for row_index, label in enumerate(labels, start=4):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=1).fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(row=row_index, column=1).font = Font(name="Aptos", bold=True, color=DARK)
        ws.cell(row=row_index, column=2).fill = PatternFill("solid", fgColor=LIGHTER)
        ws.cell(row=row_index, column=2).font = Font(name="Aptos", bold=True, color=TEXT)
        for column in (1, 2):
            ws.cell(row=row_index, column=column).border = Border(bottom=THIN_GRAY)
            ws.cell(row=row_index, column=column).alignment = Alignment(vertical="center")

    ws["B4"] = config.get("run_id") or "benchmark"
    summary_column_letters = {
        name: get_column_letter(index + 1) for index, name in enumerate(MODEL_SUMMARY_COLUMNS)
    }
    count_formulas: list[str] = []
    ok_formulas: list[str] = []
    for sheet_name, row_count, columns in result_chunks:
        last_row = max(2, row_count + 1)
        id_index = 1
        status_index = columns.index("status") + 1 if "status" in columns else None
        id_letter = get_column_letter(id_index)
        count_formulas.append(f"COUNTA('{sheet_name}'!{id_letter}2:{id_letter}{last_row})")
        if status_index is not None:
            status_letter = get_column_letter(status_index)
            ok_formulas.append(
                f'COUNTIF(\'{sheet_name}\'!{status_letter}2:{status_letter}{last_row},"ok")'
            )
    summary_last_row = max(2, len(summary_rows) + 1)
    if result_chunks:
        ws["B5"] = "=" + "+".join(count_formulas or ["0"])
        ws["B7"] = "=" + "+".join(ok_formulas or ["0"])
    else:
        total_letter = summary_column_letters["total_queries"]
        ok_letter = summary_column_letters["ok_queries"]
        ws["B5"] = f"=SUM('Model Özeti'!{total_letter}2:{total_letter}{summary_last_row})"
        ws["B7"] = f"=SUM('Model Özeti'!{ok_letter}2:{ok_letter}{summary_last_row})"
    ws["B6"] = f"=COUNTA('Model Özeti'!A2:A{summary_last_row})"
    ws["B8"] = "=B5-B7"
    for cell in ("B5", "B6", "B7", "B8"):
        ws[cell].number_format = "#,##0"

    ws.merge_cells("A10:H11")
    ws["A10"] = (
        "Ana değerlendirme: aynı model hem sorgu parçasına hem arama haritasına "
        "uygulanır. Ana başarı ölçütü tüm sorgular üzerinden 25 m içinde konumlamadır; "
        "clean ve hard_v1 koşulları ayrı raporlanır; aşağıdaki sıralama yalnız global aramayı "
        "gösterir; tüm-hata ortalaması yalnız tanısaldır."
    )
    if lightweight:
        ws["A10"].value += (
            " Bu ara raporda ham sonuçlar yoktur; tam tablo koşu sonunda eklenir."
        )
    ws["A10"].fill = PatternFill("solid", fgColor="F3F4F6")
    ws["A10"].font = Font(name="Aptos", italic=True, color="4B5563")
    ws["A10"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A10"].border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
    ws.row_dimensions[10].height = 26
    ws.row_dimensions[11].height = 18

    ranking_headers = [
        "Sıra",
        "Model",
        "Senaryo",
        "Yön",
        "Başarı ≤25 m",
        "AUC@25 m",
        "Medyan hata (m)",
    ]
    for column, header in enumerate(ranking_headers, start=1):
        ws.cell(row=14, column=column, value=header)
    style_header(ws[14][: len(ranking_headers)])
    ws.row_dimensions[14].height = 32

    summary_positions = sorted(
        [
            row
            for row in summary_rows
            if str(row.get("search_mode") or "").lower() == "global"
            and row.get("success_25m") is not None
        ],
        key=lambda item: (
            -float(item.get("success_25m") or 0.0),
            -float(item.get("auc_25m") or 0.0),
            (
                float(item["median_error_under_25m"])
                if item.get("median_error_under_25m") is not None
                else float("inf")
            ),
        ),
    )[:10]
    chart_positions = summary_positions[:5]
    for offset, row in enumerate(summary_positions, start=15):
        values = [
            offset - 14,
            compact_model_label(row),
            display_query_variant(row.get("query_variant")),
            display_direction(row.get("direction")),
            row.get("success_25m"),
            row.get("auc_25m"),
            row.get("median_error_under_25m"),
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row=offset, column=column, value=excel_value(value))
    for row in range(15, 25):
        for column in range(1, len(ranking_headers) + 1):
            ws.cell(row=row, column=column).border = SUBTLE_BORDER
            ws.cell(row=row, column=column).font = Font(name="Aptos", size=9, color=TEXT)
        ws.cell(row=row, column=5).number_format = "0.0%"
        ws.cell(row=row, column=6).number_format = "0.0%"
        ws.cell(row=row, column=7).number_format = "0.00"

    if chart_positions:
        # Keep a compact, formula-backed Top-5 block beside the ranking table.
        # Values are converted to 0-100 for reliable percent-axis rendering.
        # This block is rebuilt on every full, lightweight and incremental save.
        ws["I13"] = "Grafik: En iyi 5"
        ws["I13"].font = Font(name="Aptos", size=10, bold=True, color=TEXT)
        ws["I14"] = "Model | Senaryo"
        ws["J14"] = "Başarı (%)"
        for cell in ws["I14:J14"][0]:
            cell.fill = PatternFill("solid", fgColor=DARK)
            cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row, item in enumerate(chart_positions, start=15):
            ws.cell(
                row=row,
                column=9,
                value=(
                    f"{compact_model_label(item)} | "
                    f"{display_query_variant(item.get('query_variant'))} | "
                    f"{compact_direction_label(item.get('direction'))}"
                ),
            )
            ws.cell(row=row, column=10, value=f"=E{row}*100")
            for column in range(9, 11):
                cell = ws.cell(row=row, column=column)
                cell.font = Font(name="Aptos", size=9, color=TEXT)
                cell.border = SUBTLE_BORDER
                cell.alignment = Alignment(vertical="center")
            ws.cell(row=row, column=10).number_format = "0.0"
        ws.column_dimensions["I"].width = 22
        ws.column_dimensions["J"].width = 12
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "En iyi 5 model | 25 m başarı (%)"
        chart.height = 10.0
        chart.width = 15.0
        chart.varyColors = False
        data = Reference(ws, min_col=10, min_row=14, max_row=14 + len(chart_positions))
        categories = Reference(ws, min_col=9, min_row=15, max_row=14 + len(chart_positions))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend = None
        chart.y_axis.title = "Başarı (%)"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.y_axis.numFmt = "0"
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = ACCENT
            chart.series[0].graphicalProperties.line.solidFill = ACCENT
        ws.add_chart(chart, "L14")
    widths = {"A": 8, "B": 14, "C": 12, "D": 38, "E": 14, "F": 12, "G": 18, "H": 3}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for column in range(11, 19):
        ws.column_dimensions[get_column_letter(column)].width = 12
    ws.print_area = "A1:R30"
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def locked_copy_path(output_path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = output_path.with_name(f"{output_path.stem}_{stamp}_kilitli_kopya{output_path.suffix}")
    if not base.exists():
        return base
    for index in range(2, 1000):
        candidate = output_path.with_name(
            f"{output_path.stem}_{stamp}_kilitli_kopya_{index:02d}{output_path.suffix}"
        )
        if not candidate.exists():
            return candidate
    raise RuntimeError("Kilitli Excel için benzersiz kopya adı üretilemedi.")


def save_workbook_safely(wb: Workbook, output_path: Path) -> tuple[Path, str]:
    """Save atomically; if Excel locks the target, preserve it and publish a copy."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(
        f".{output_path.name}.{os.getpid()}.{datetime.now().strftime('%Y%m%d%H%M%S%f')}.tmp"
    )
    LOG.info("Çalışma kitabı geçici dosyaya yazılıyor: %s", temporary)
    try:
        wb.save(temporary)
        try:
            os.replace(temporary, output_path)
            return output_path, "primary"
        except PermissionError:
            if not output_path.exists():
                raise
            alternate = locked_copy_path(output_path)
            os.replace(temporary, alternate)
            LOG.warning(
                "ANA EXCEL KİLİTLİ | benchmark devam ediyor | güncel rapor kopyaya yazıldı: %s",
                alternate,
            )
            return alternate, "locked_copy"
    finally:
        if temporary.exists():
            try:
                temporary.unlink()
            except OSError:
                LOG.warning("Geçici Excel dosyası silinemedi: %s", temporary)


def write_latest_pointer(
    run_dir: Path,
    *,
    requested_path: Path,
    actual_path: Path,
    output_mode: str,
) -> None:
    pointer_path = run_dir / LATEST_POINTER
    temporary = pointer_path.with_suffix(pointer_path.suffix + f".{os.getpid()}.tmp")
    payload = {
        "updated_at_utc": utc_now_iso(),
        "requested_workbook": str(requested_path.resolve()),
        "latest_workbook": str(actual_path.resolve()),
        "output_mode": output_mode,
    }
    try:
        temporary.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        os.replace(temporary, pointer_path)
    except OSError:
        LOG.exception("Excel son-kopya işaretçisi yazılamadı; benchmark devam edecek.")
    finally:
        if temporary.exists():
            try:
                temporary.unlink()
            except OSError:
                pass


def resolve_incremental_base(run_dir: Path, output_path: Path) -> Path:
    candidates: list[Path] = []
    resolved_output = output_path.resolve()
    if resolved_output.is_file():
        candidates.append(resolved_output)
    pointer = read_json(run_dir / LATEST_POINTER, {})
    latest_text = pointer.get("latest_workbook") if isinstance(pointer, dict) else None
    if isinstance(latest_text, str):
        latest = Path(latest_text).resolve()
        if (
            latest.is_file()
            and latest.suffix.lower() == ".xlsx"
            and latest.parent == resolved_output.parent
        ):
            candidates.append(latest)
    if not candidates:
        return output_path
    return max(candidates, key=lambda path: path.stat().st_mtime_ns)


def build_workbook(
    run_dir: Path,
    output_path: Path,
    *,
    validation_mode: str = "deep",
    lightweight: bool = False,
) -> dict[str, Any]:
    LOG.info("Kaynak dosyalar okunuyor: %s", run_dir)
    config = read_json(run_dir / "run_config.json", {})
    summary_rows = read_json(run_dir / "summary.json", [])
    result_rows = [] if lightweight else read_jsonl(run_dir / "results.jsonl")
    error_rows = read_jsonl(run_dir / "model_errors.jsonl")
    query_rows = [] if lightweight else find_query_rows(run_dir)
    if not isinstance(summary_rows, list):
        raise ValueError("summary.json bir liste içermelidir.")
    summary_rows = enrich_summary_model_identity(run_dir, summary_rows)
    summary_display_rows = prepare_model_summary_rows(summary_rows)
    model_catalog_rows = build_model_catalog_rows(run_dir, summary_rows)
    LOG.info(
        "Kayıt sayıları | sonuç=%s | özet=%d | sorgu=%s | model_hatası=%d | kapsam=%s",
        len(result_rows) if not lightweight else "atlanıyor",
        len(summary_rows),
        len(query_rows) if not lightweight else "atlanıyor",
        len(error_rows),
        "hafif" if lightweight else "tam",
    )

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Özet")
    summary_ws = wb.create_sheet("Model Özeti")
    catalog_ws = wb.create_sheet("Model Kataloğu")
    raw_ws = wb.create_sheet("Ham Sonuçlar") if not lightweight else None
    query_ws = wb.create_sheet("Sorgu Manifesti") if not lightweight else None
    config_ws = wb.create_sheet("Yapılandırma")
    error_ws = wb.create_sheet("Hatalar")
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass

    summary_columns = MODEL_SUMMARY_COLUMNS
    LOG.info("Model Özeti sayfası yazılıyor (%d satır, %d görünür alan).", len(summary_display_rows), len(summary_columns) - 1)
    write_table(
        summary_ws,
        summary_display_rows,
        summary_columns,
        table_name="ModelSummary_01",
        number_formats=SUMMARY_FORMATS,
        headers=[MODEL_SUMMARY_HEADERS[column] for column in summary_columns],
    )
    set_widths(
        summary_ws,
        summary_columns,
        overrides={
            "model_sequence": 10,
            "direction": 22,
            "query_variant": 12,
            "search_mode": 14,
            "model_label": 28,
            "model_id": 48,
            "model_epoch": 10,
            "total_queries": 10,
            "ok_queries": 12,
            "coverage": 11,
            "success_25m": 14,
            "auc_25m": 12,
            "median_error_under_25m": 20,
            "p90_error_m": 14,
            "mean_search_seconds": 16,
        },
    )
    summary_ws.column_dimensions[
        get_column_letter(summary_columns.index("ok_queries") + 1)
    ].hidden = True
    if summary_ws.max_row >= 2 and "coverage" in summary_columns:
        coverage_letter = get_column_letter(summary_columns.index("coverage") + 1)
        summary_ws.conditional_formatting.add(
            f"{coverage_letter}2:{coverage_letter}{summary_ws.max_row}",
            CellIsRule(operator="lessThan", formula=["1"], fill=PatternFill("solid", fgColor=BAD)),
        )
    if summary_ws.max_row >= 2 and "success_25m" in summary_columns:
        success_letter = get_column_letter(summary_columns.index("success_25m") + 1)
        summary_ws.conditional_formatting.add(
            f"{success_letter}2:{success_letter}{summary_ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FECACA",
                mid_type="percentile",
                mid_value=50,
                mid_color="FEF3C7",
                end_type="max",
                end_color="BBF7D0",
            ),
        )

    LOG.info("Model Kataloğu sayfası yazılıyor (%d model).", len(model_catalog_rows))
    write_table(
        catalog_ws,
        model_catalog_rows,
        MODEL_CATALOG_COLUMNS,
        table_name="ModelCatalog_01",
        number_formats={"model_epoch": "0"},
        headers=[MODEL_CATALOG_HEADERS[column] for column in MODEL_CATALOG_COLUMNS],
    )
    set_widths(
        catalog_ws,
        MODEL_CATALOG_COLUMNS,
        overrides={
            "model_id": 48,
            "model_label": 28,
            "model_epoch": 10,
            "model_relative_path": 60,
            "model_sha256": 68,
        },
    )

    result_columns = ordered_columns(result_rows, ["run_id", "status"])
    raw_chunks: list[tuple[str, int, Sequence[str]]] = []
    if not lightweight:
        assert raw_ws is not None and query_ws is not None
        per_sheet = MAX_EXCEL_ROWS - 1
        chunks = [
            result_rows[index : index + per_sheet]
            for index in range(0, len(result_rows), per_sheet)
        ]
        if not chunks:
            chunks = [[]]
        for chunk_index, chunk in enumerate(chunks, start=1):
            if chunk_index == 1:
                ws = raw_ws
                sheet_name = "Ham Sonuçlar"
            else:
                sheet_name = f"Ham Sonuçlar {chunk_index}"
                ws = wb.create_sheet(sheet_name)
            LOG.info("%s sayfası yazılıyor (%d satır).", sheet_name, len(chunk))
            write_table(
                ws,
                chunk,
                result_columns,
                table_name=safe_table_name("RawResults", chunk_index),
                number_formats=RESULT_FORMATS,
            )
            set_widths(
                ws,
                result_columns,
                overrides={
                    "model_id": 48,
                    "model_file": 56,
                    "source_query_raster": 48,
                    "source_map_raster": 48,
                    "reason": 38,
                },
            )
            add_error_conditional_formatting(ws, result_columns)
            raw_chunks.append((sheet_name, len(chunk), result_columns))

        query_columns = ordered_columns(query_rows, QUERY_COLUMNS)
        LOG.info("Sorgu Manifesti sayfası yazılıyor (%d satır).", len(query_rows))
        write_table(
            query_ws,
            query_rows,
            query_columns,
            table_name="QueryManifest_01",
            number_formats=QUERY_FORMATS,
        )
        set_widths(
            query_ws,
            query_columns,
            overrides={"direction": 38, "raw_tile_file": 58},
        )

    config_rows = [
        {"Parametre": key, "Değer": excel_value(value)} for key, value in sorted(config.items())
    ]
    LOG.info("Yapılandırma sayfası yazılıyor (%d parametre).", len(config_rows))
    write_table(
        config_ws,
        config_rows,
        ["Parametre", "Değer"],
        table_name="RunConfig_01",
    )
    set_widths(config_ws, ["Parametre", "Değer"], overrides={"Parametre": 34, "Değer": 90})
    for cell in config_ws["B"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    error_columns = ordered_columns(
        error_rows, ["created_at_utc", "direction", "model_id", "error_type", "error"]
    )
    LOG.info("Hatalar sayfası yazılıyor (%d satır).", len(error_rows))
    write_table(
        error_ws,
        error_rows,
        error_columns,
        table_name="ModelErrors_01",
    )
    set_widths(
        error_ws,
        error_columns,
        overrides={"model_file": 58, "error": 62, "traceback": 70},
    )
    for name in ("error", "traceback"):
        if name in error_columns:
            index = error_columns.index(name) + 1
            for row in range(2, error_ws.max_row + 1):
                error_ws.cell(row=row, column=index).alignment = Alignment(vertical="top", wrap_text=True)

    LOG.info("Özet panosu ve formül bağlı grafik hazırlanıyor.")
    write_dashboard(wb, config, summary_display_rows, raw_chunks, lightweight=lightweight)
    wb.active = wb.sheetnames.index("Özet")

    if not lightweight:
        write_excel_state(
            wb,
            build_excel_state(
                run_dir,
                result_rows=len(result_rows),
                result_columns=result_columns,
            ),
        )
    actual_path, output_mode = save_workbook_safely(wb, output_path)
    LOG.info(
        "Çalışma kitabı kaydedildi: %s (%.2f MiB)",
        actual_path,
        actual_path.stat().st_size / 2**20,
    )
    report = validate_workbook(
        actual_path,
        run_dir,
        mode=validation_mode,
        requested_path=output_path,
        output_mode=output_mode,
        lightweight=lightweight,
    )
    write_latest_pointer(
        run_dir,
        requested_path=output_path,
        actual_path=actual_path,
        output_mode=output_mode,
    )
    return report


def replace_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        index = wb.sheetnames.index(name)
        wb.remove(wb[name])
    else:
        index = len(wb.sheetnames)
    return wb.create_sheet(name, index)


def raw_table_layout(wb: Workbook) -> tuple[Any, Any, list[str], int]:
    extra_raw = [name for name in wb.sheetnames if name.startswith("Ham Sonuçlar ")]
    if extra_raw:
        raise IncrementalUpdateUnavailable(
            "Birden fazla ham-sonuç sayfası var; güvenli artımlı ekleme yerine tam üretim gerekir."
        )
    if "Ham Sonuçlar" not in wb.sheetnames:
        raise IncrementalUpdateUnavailable("Ham Sonuçlar sayfası bulunamadı.")
    ws = wb["Ham Sonuçlar"]
    tables = list(ws.tables.values())
    if len(tables) != 1:
        raise IncrementalUpdateUnavailable("Ham Sonuçlar sayfasında tek tablo bulunmalıdır.")
    table = tables[0]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if min_col != 1 or min_row != 1:
        raise IncrementalUpdateUnavailable("Ham sonuç tablosu A1 hücresinden başlamıyor.")
    columns = [ws.cell(row=1, column=index).value for index in range(1, max_col + 1)]
    if not all(isinstance(value, str) and value for value in columns):
        raise IncrementalUpdateUnavailable("Ham sonuç sütun başlıkları geçersiz.")
    existing_count = max(0, max_row - 1)
    if existing_count == 1 and all(
        ws.cell(row=2, column=index).value is None for index in range(1, max_col + 1)
    ):
        existing_count = 0
    return ws, table, list(columns), existing_count


def parse_jsonl_record(raw_line: bytes, path: Path, line_number: int) -> dict[str, Any]:
    try:
        value = json.loads(raw_line.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise IncrementalUpdateUnavailable(
            f"{path}:{line_number} geçerli UTF-8 JSONL kaydı değil."
        ) from exc
    if not isinstance(value, dict):
        raise IncrementalUpdateUnavailable(f"{path}:{line_number} bir JSON nesnesi değil.")
    return value


def read_incremental_result_rows(
    path: Path,
    *,
    ws,
    columns: Sequence[str],
    existing_count: int,
    state: dict[str, Any] | None,
) -> tuple[list[dict[str, Any]], dict[str, Any], bool]:
    """Return new append-only rows, a source identity, and whether legacy bootstrap ran."""
    if not path.is_file():
        if existing_count:
            raise IncrementalUpdateUnavailable("results.jsonl kayıp, fakat Excel'de ham satırlar var.")
        empty_digest = hashlib.sha256().hexdigest()
        return [], {"exists": False, "size_bytes": 0, "sha256": empty_digest}, state is None

    snapshot_size = path.stat().st_size
    digest = hashlib.sha256()
    new_rows: list[dict[str, Any]] = []
    column_set = set(columns)
    bootstrapped = state is None

    with path.open("rb") as handle:
        if state is not None:
            if state.get("schema_version") != STATE_SCHEMA_VERSION:
                raise IncrementalUpdateUnavailable("Excel artımlı durum sürümü uyumsuz.")
            if int(state.get("result_rows", -1)) != existing_count:
                raise IncrementalUpdateUnavailable("Excel durum satır sayısı tabloyla uyuşmuyor.")
            if list(state.get("result_columns") or []) != list(columns):
                raise IncrementalUpdateUnavailable("Excel durum sütunları tablo başlıklarıyla uyuşmuyor.")
            previous = state.get("results") or {}
            previous_size = int(previous.get("size_bytes", -1))
            if previous_size < 0 or previous_size > snapshot_size:
                raise IncrementalUpdateUnavailable("results.jsonl küçülmüş veya durum ofseti geçersiz.")
            remaining = previous_size
            while remaining:
                block = handle.read(min(8 * 1024 * 1024, remaining))
                if not block:
                    raise IncrementalUpdateUnavailable("results.jsonl beklenenden erken bitti.")
                digest.update(block)
                remaining -= len(block)
            if digest.hexdigest() != previous.get("sha256"):
                raise IncrementalUpdateUnavailable(
                    "results.jsonl önceki Excel checkpointinden sonra geriye dönük değiştirilmiş."
                )
            line_number = existing_count
            remaining = snapshot_size - previous_size
            while remaining:
                raw_line = handle.readline(remaining)
                if not raw_line:
                    raise IncrementalUpdateUnavailable("results.jsonl ek bölümü beklenenden erken bitti.")
                remaining -= len(raw_line)
                digest.update(raw_line)
                line_number += 1
                if not raw_line.endswith(b"\n"):
                    raise IncrementalUpdateUnavailable("results.jsonl son kaydı tamamlanmamış.")
                if not raw_line.strip():
                    continue
                row = parse_jsonl_record(raw_line, path, line_number)
                unknown = set(row) - column_set
                if unknown:
                    raise IncrementalUpdateUnavailable(
                        f"Yeni ham sonuç sütunları bulundu: {sorted(unknown)}"
                    )
                new_rows.append(row)
        else:
            source_row_count = 0
            remaining = snapshot_size
            while remaining:
                raw_line = handle.readline(remaining)
                if not raw_line:
                    raise IncrementalUpdateUnavailable("results.jsonl beklenenden erken bitti.")
                remaining -= len(raw_line)
                digest.update(raw_line)
                if not raw_line.endswith(b"\n"):
                    raise IncrementalUpdateUnavailable("results.jsonl son kaydı tamamlanmamış.")
                if not raw_line.strip():
                    continue
                source_row_count += 1
                row = parse_jsonl_record(raw_line, path, source_row_count)
                unknown = set(row) - column_set
                if unknown:
                    raise IncrementalUpdateUnavailable(
                        f"Eski Excel'de bulunmayan ham sonuç sütunları var: {sorted(unknown)}"
                    )
                if source_row_count <= existing_count:
                    excel_row = source_row_count + 1
                    for column_index, column in enumerate(columns, start=1):
                        expected = excel_value(row.get(column))
                        actual = ws.cell(row=excel_row, column=column_index).value
                        if not excel_values_equal(actual, expected):
                            raise IncrementalUpdateUnavailable(
                                "Eski Excel ham verisi kaynak JSONL önekiyle birebir uyuşmuyor "
                                f"({ws.title}!{get_column_letter(column_index)}{excel_row})."
                            )
                else:
                    new_rows.append(row)
            if source_row_count < existing_count:
                raise IncrementalUpdateUnavailable(
                    "Excel'deki ham satır sayısı results.jsonl kayıt sayısından fazla."
                )

    identity = {"exists": True, "size_bytes": snapshot_size, "sha256": digest.hexdigest()}
    return new_rows, identity, bootstrapped


def append_raw_rows(
    ws,
    table,
    columns: Sequence[str],
    rows: Sequence[dict[str, Any]],
    *,
    existing_count: int,
) -> int:
    if not rows:
        return existing_count
    if existing_count == 0 and ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    start_row = existing_count + 2
    for row in rows:
        ws.append([excel_value(row.get(column)) for column in columns])
    end_row = existing_count + len(rows) + 1
    style_data_rows(
        ws,
        columns,
        min_row=start_row,
        max_row=end_row,
        number_formats=RESULT_FORMATS,
    )
    table.ref = f"A1:{get_column_letter(len(columns))}{end_row}"
    ws.conditional_formatting = ConditionalFormattingList()
    add_error_conditional_formatting(ws, columns, min_row=2, max_row=end_row)
    return existing_count + len(rows)


def rebuild_small_sheets(
    wb: Workbook,
    run_dir: Path,
    *,
    config: dict[str, Any],
    summary_rows: Sequence[dict[str, Any]],
    result_count: int,
    result_columns: Sequence[str],
) -> None:
    summary_display_rows = prepare_model_summary_rows(summary_rows)
    model_catalog_rows = build_model_catalog_rows(run_dir, summary_rows)
    summary_ws = replace_sheet(wb, "Model Özeti")
    summary_columns = MODEL_SUMMARY_COLUMNS
    write_table(
        summary_ws,
        summary_display_rows,
        summary_columns,
        table_name="ModelSummary_01",
        number_formats=SUMMARY_FORMATS,
        headers=[MODEL_SUMMARY_HEADERS[column] for column in summary_columns],
    )
    set_widths(
        summary_ws,
        summary_columns,
        overrides={
            "model_sequence": 10,
            "direction": 22,
            "query_variant": 12,
            "search_mode": 14,
            "model_label": 28,
            "model_id": 48,
            "model_epoch": 10,
            "total_queries": 10,
            "ok_queries": 12,
            "coverage": 11,
            "success_25m": 14,
            "auc_25m": 12,
            "median_error_under_25m": 20,
            "p90_error_m": 14,
            "mean_search_seconds": 16,
        },
    )
    summary_ws.column_dimensions[
        get_column_letter(summary_columns.index("ok_queries") + 1)
    ].hidden = True
    if summary_ws.max_row >= 2 and "coverage" in summary_columns:
        letter = get_column_letter(summary_columns.index("coverage") + 1)
        summary_ws.conditional_formatting.add(
            f"{letter}2:{letter}{summary_ws.max_row}",
            CellIsRule(operator="lessThan", formula=["1"], fill=PatternFill("solid", fgColor=BAD)),
        )
    if summary_ws.max_row >= 2 and "success_25m" in summary_columns:
        letter = get_column_letter(summary_columns.index("success_25m") + 1)
        summary_ws.conditional_formatting.add(
            f"{letter}2:{letter}{summary_ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FECACA",
                mid_type="percentile",
                mid_value=50,
                mid_color="FEF3C7",
                end_type="max",
                end_color="BBF7D0",
            ),
        )

    catalog_ws = replace_sheet(wb, "Model Kataloğu")
    write_table(
        catalog_ws,
        model_catalog_rows,
        MODEL_CATALOG_COLUMNS,
        table_name="ModelCatalog_01",
        number_formats={"model_epoch": "0"},
        headers=[MODEL_CATALOG_HEADERS[column] for column in MODEL_CATALOG_COLUMNS],
    )
    set_widths(
        catalog_ws,
        MODEL_CATALOG_COLUMNS,
        overrides={
            "model_id": 48,
            "model_label": 28,
            "model_epoch": 10,
            "model_relative_path": 60,
            "model_sha256": 68,
        },
    )

    query_rows = find_query_rows(run_dir)
    query_columns = ordered_columns(query_rows, QUERY_COLUMNS)
    query_ws = replace_sheet(wb, "Sorgu Manifesti")
    write_table(
        query_ws,
        query_rows,
        query_columns,
        table_name="QueryManifest_01",
        number_formats=QUERY_FORMATS,
    )
    set_widths(query_ws, query_columns, overrides={"direction": 38, "raw_tile_file": 58})

    config_rows = [
        {"Parametre": key, "Değer": excel_value(value)} for key, value in sorted(config.items())
    ]
    config_ws = replace_sheet(wb, "Yapılandırma")
    write_table(config_ws, config_rows, ["Parametre", "Değer"], table_name="RunConfig_01")
    set_widths(config_ws, ["Parametre", "Değer"], overrides={"Parametre": 34, "Değer": 90})
    for cell in config_ws["B"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    error_rows = read_jsonl(run_dir / "model_errors.jsonl")
    error_columns = ordered_columns(
        error_rows, ["created_at_utc", "direction", "model_id", "error_type", "error"]
    )
    error_ws = replace_sheet(wb, "Hatalar")
    write_table(error_ws, error_rows, error_columns, table_name="ModelErrors_01")
    set_widths(
        error_ws,
        error_columns,
        overrides={"model_file": 58, "error": 62, "traceback": 70},
    )
    for name in ("error", "traceback"):
        if name in error_columns:
            index = error_columns.index(name) + 1
            for row in range(2, error_ws.max_row + 1):
                error_ws.cell(row=row, column=index).alignment = Alignment(
                    vertical="top", wrap_text=True
                )

    replace_sheet(wb, "Özet")
    write_dashboard(
        wb,
        config,
        summary_display_rows,
        [("Ham Sonuçlar", result_count, result_columns)],
    )
    wb.active = wb.sheetnames.index("Özet")


def update_workbook_incremental(
    run_dir: Path,
    output_path: Path,
    *,
    validation_mode: str = "checkpoint",
) -> dict[str, Any]:
    base_path = resolve_incremental_base(run_dir, output_path)
    if not base_path.is_file():
        raise IncrementalUpdateUnavailable("Artımlı güncellenecek mevcut Excel bulunamadı.")
    LOG.info("Artımlı Excel tabanı yükleniyor: %s", base_path)
    wb = load_workbook(base_path, data_only=False, read_only=False)
    missing = [name for name in REQUIRED_SHEETS if name not in wb.sheetnames]
    if missing:
        raise IncrementalUpdateUnavailable(f"Mevcut Excel'de zorunlu sayfalar eksik: {missing}")

    raw_ws, raw_table, result_columns, existing_count = raw_table_layout(wb)
    state = read_excel_state(wb)
    new_rows, results_identity, bootstrapped = read_incremental_result_rows(
        run_dir / "results.jsonl",
        ws=raw_ws,
        columns=result_columns,
        existing_count=existing_count,
        state=state,
    )
    total_count = append_raw_rows(
        raw_ws,
        raw_table,
        result_columns,
        new_rows,
        existing_count=existing_count,
    )
    LOG.info(
        "ARTIMLI HAM SONUÇ | önceki=%d | yeni=%d | toplam=%d | eski_rapor_doğrulama=%s",
        existing_count,
        len(new_rows),
        total_count,
        "evet" if bootstrapped else "hayır",
    )

    config = read_json(run_dir / "run_config.json", {})
    summary_rows = read_json(run_dir / "summary.json", [])
    if not isinstance(config, dict):
        raise IncrementalUpdateUnavailable("run_config.json bir nesne içermelidir.")
    if not isinstance(summary_rows, list):
        raise IncrementalUpdateUnavailable("summary.json bir liste içermelidir.")
    summary_rows = enrich_summary_model_identity(run_dir, summary_rows)
    rebuild_small_sheets(
        wb,
        run_dir,
        config=config,
        summary_rows=summary_rows,
        result_count=total_count,
        result_columns=result_columns,
    )
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass
    write_excel_state(
        wb,
        build_excel_state(
            run_dir,
            result_rows=total_count,
            result_columns=result_columns,
            results_identity=results_identity,
        ),
    )
    actual_path, output_mode = save_workbook_safely(wb, output_path)
    report = validate_workbook(
        actual_path,
        run_dir,
        mode=validation_mode,
        requested_path=output_path,
        output_mode=output_mode,
    )
    report["incremental"] = True
    report["appended_result_rows"] = len(new_rows)
    report["base_workbook"] = str(base_path.resolve())
    write_latest_pointer(
        run_dir,
        requested_path=output_path,
        actual_path=actual_path,
        output_mode=output_mode,
    )
    return report


def validate_workbook_checkpoint(
    output_path: Path,
    run_dir: Path,
    *,
    requested_path: Path,
    output_mode: str,
    lightweight: bool = False,
) -> dict[str, Any]:
    """Fast model-boundary validation without reloading millions of cells."""
    LOG.info("XLSX checkpoint ZIP, sayfa, tablo, formül ve grafik yapısı doğrulanıyor.")
    with zipfile.ZipFile(output_path) as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise RuntimeError(f"XLSX arşiv üyesi bozuk: {bad_member}")

        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        sheets = []
        overview_rid = None
        state_hidden = False
        for element in workbook_root.findall(f".//{{{main_ns}}}sheet"):
            name = element.attrib.get("name", "")
            sheets.append(name)
            if name == "Özet":
                overview_rid = element.attrib.get(f"{{{rel_ns}}}id")
            if name == STATE_SHEET:
                state_hidden = element.attrib.get("state") in {"hidden", "veryHidden"}
        required_sheets = LIGHTWEIGHT_SHEETS if lightweight else REQUIRED_SHEETS
        missing = [name for name in required_sheets if name not in sheets]
        if missing:
            raise RuntimeError(f"Eksik çalışma sayfaları: {missing}")
        if not lightweight and (STATE_SHEET not in sheets or not state_hidden):
            raise RuntimeError("Artımlı Excel durum sayfası eksik veya gizli değil.")

        table_names: list[str] = []
        table_refs: dict[str, str] = {}
        for member in archive.namelist():
            if not member.startswith("xl/tables/table") or not member.endswith(".xml"):
                continue
            root = ElementTree.fromstring(archive.read(member))
            name = root.attrib.get("displayName") or root.attrib.get("name") or member
            table_names.append(name)
            table_refs[name] = root.attrib.get("ref", "")
        expected_tables = {"ModelSummary_01", "ModelCatalog_01", "RunConfig_01", "ModelErrors_01"}
        if not lightweight:
            expected_tables.update({"RawResults_01", "QueryManifest_01"})
        missing_tables = sorted(expected_tables - set(table_names))
        if missing_tables:
            raise RuntimeError(f"Eksik Excel tabloları: {missing_tables}")

        formula_count = 0
        if overview_rid:
            rel_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            package_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
            target = None
            for relationship in rel_root.findall(f"{{{package_ns}}}Relationship"):
                if relationship.attrib.get("Id") == overview_rid:
                    target = relationship.attrib.get("Target")
                    break
            if target:
                normalized = target.lstrip("/")
                if not normalized.startswith("xl/"):
                    normalized = f"xl/{normalized}"
                overview_root = ElementTree.fromstring(archive.read(normalized))
                formula_count = sum(1 for _ in overview_root.iter(f"{{{main_ns}}}f"))
        if formula_count <= 0:
            raise RuntimeError("Özet panosunda formül bulunamadı.")
        chart_count = sum(
            1
            for member in archive.namelist()
            if member.startswith("xl/charts/chart") and member.endswith(".xml")
        )
        if chart_count <= 0:
            raise RuntimeError("Özet grafiği bulunamadı.")

    sha256 = sha256_file(output_path)
    report = {
        "validated_at_utc": utc_now_iso(),
        "validation_mode": "checkpoint",
        "report_scope": "lightweight" if lightweight else "full",
        "engine": "openpyxl",
        "openpyxl_version": __import__("openpyxl").__version__,
        "requested_workbook": str(requested_path.resolve()),
        "workbook": str(output_path.resolve()),
        "output_mode": output_mode,
        "size_bytes": output_path.stat().st_size,
        "sha256": sha256,
        "sheet_count": len(sheets),
        "sheet_names": sheets,
        "tables": table_refs,
        "formula_count": formula_count,
        "formula_error_literals": [],
        "formula_error_scan": "deferred_to_final_deep_validation",
        "chart_count": chart_count,
        "zip_integrity": "ok",
    }
    validation_path = run_dir / "excel_validation.json"
    validation_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    LOG.info(
        "XLSX checkpoint doğrulandı | sayfa=%d | formül=%d | grafik=%d | sha256=%s",
        len(sheets),
        formula_count,
        chart_count,
        sha256[:16],
    )
    return report


def validate_workbook(
    output_path: Path,
    run_dir: Path,
    *,
    mode: str = "deep",
    requested_path: Path | None = None,
    output_mode: str = "primary",
    lightweight: bool = False,
) -> dict[str, Any]:
    requested_path = requested_path or output_path
    if mode == "checkpoint":
        return validate_workbook_checkpoint(
            output_path,
            run_dir,
            requested_path=requested_path,
            output_mode=output_mode,
            lightweight=lightweight,
        )
    if mode != "deep":
        raise ValueError(f"Bilinmeyen Excel doğrulama modu: {mode}")
    LOG.info("XLSX ZIP bütünlüğü ve sayfa yapısı doğrulanıyor.")
    with zipfile.ZipFile(output_path) as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise RuntimeError(f"XLSX arşiv üyesi bozuk: {bad_member}")

    wb = load_workbook(output_path, data_only=False, read_only=False)
    required_sheets = LIGHTWEIGHT_SHEETS if lightweight else REQUIRED_SHEETS
    missing = [name for name in required_sheets if name not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Eksik çalışma sayfaları: {missing}")

    formula_count = 0
    formula_error_literals: list[str] = []
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                elif isinstance(value, str) and value in {
                    "#REF!",
                    "#DIV/0!",
                    "#VALUE!",
                    "#NAME?",
                    "#N/A",
                }:
                    formula_error_literals.append(f"{ws.title}!{cell.coordinate}:{value}")
        sheets.append(
            {
                "name": ws.title,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "tables": sorted(ws.tables.keys()),
                "charts": len(ws._charts),
            }
        )
    if formula_error_literals:
        raise RuntimeError(f"Formül hata sabitleri bulundu: {formula_error_literals[:10]}")

    sha256 = hashlib.sha256(output_path.read_bytes()).hexdigest()
    report = {
        "validated_at_utc": utc_now_iso(),
        "validation_mode": "deep",
        "report_scope": "lightweight" if lightweight else "full",
        "engine": "openpyxl",
        "openpyxl_version": __import__("openpyxl").__version__,
        "requested_workbook": str(requested_path.resolve()),
        "workbook": str(output_path.resolve()),
        "output_mode": output_mode,
        "size_bytes": output_path.stat().st_size,
        "sha256": sha256,
        "sheet_count": len(wb.sheetnames),
        "sheets": sheets,
        "formula_count": formula_count,
        "formula_error_literals": formula_error_literals,
        "zip_integrity": "ok",
    }
    validation_path = run_dir / "excel_validation.json"
    validation_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    LOG.info(
        "XLSX doğrulandı | sayfa=%d | formül=%d | grafik=%d | sha256=%s",
        len(wb.sheetnames),
        formula_count,
        sum(item["charts"] for item in sheets),
        sha256[:16],
    )
    LOG.info("Doğrulama kaydı: %s", validation_path)
    return report


def main(argv: Sequence[str] | None = None) -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | EXCEL | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )
    args = parse_args(argv)
    run_dir = args.run_dir.resolve()
    output_path = (args.output or (run_dir / "benchmark_results.xlsx")).resolve()
    if not run_dir.is_dir():
        raise FileNotFoundError(f"Koşu klasörü bulunamadı: {run_dir}")
    if args.incremental and args.lightweight:
        raise ValueError("--incremental ve --lightweight birlikte kullanılamaz.")
    if args.lightweight:
        report = build_workbook(
            run_dir,
            output_path,
            validation_mode=args.validation_mode,
            lightweight=True,
        )
    elif args.incremental:
        try:
            report = update_workbook_incremental(
                run_dir,
                output_path,
                validation_mode=args.validation_mode,
            )
        except IncrementalUpdateUnavailable as exc:
            LOG.warning(
                "ARTIMLI EXCEL GÜVENLİ DEĞİL | tam üretime dönülüyor | neden=%s",
                exc,
            )
            report = build_workbook(
                run_dir,
                output_path,
                validation_mode=args.validation_mode,
            )
            report["incremental"] = False
            report["incremental_fallback_reason"] = str(exc)
    else:
        report = build_workbook(
            run_dir,
            output_path,
            validation_mode=args.validation_mode,
        )
    actual_path = Path(report["workbook"])
    print("WORKBOOK_READY_JSON: " + json.dumps({"path": str(actual_path), "report": report}))
    print(f"Workbook ready: {actual_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
