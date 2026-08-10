import json
import ctypes
import sys
import tempfile
import time
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl import load_workbook


BENCHMARK_DIR = Path(__file__).resolve().parents[1]
if str(BENCHMARK_DIR) not in sys.path:
    sys.path.insert(0, str(BENCHMARK_DIR))

from build_benchmark_excel_openpyxl import (  # noqa: E402
    SUMMARY_COLUMNS,
    build_workbook,
    excel_values_equal,
    resolve_incremental_base,
    save_workbook_safely,
    update_workbook_incremental,
)


class OpenpyxlReportTests(unittest.TestCase):
    def test_excel_empty_cell_and_json_empty_string_are_losslessly_equivalent(self):
        self.assertTrue(excel_values_equal(None, ""))
        self.assertTrue(excel_values_equal("", None))
        self.assertFalse(excel_values_equal(None, 0))
        self.assertTrue(excel_values_equal(0.001250270328719723, 0.0012502703287197233))
        self.assertFalse(excel_values_equal(1.0, 1.000001))

    def test_newer_primary_wins_over_a_stale_locked_copy_pointer(self):
        with tempfile.TemporaryDirectory() as temporary:
            run_dir = Path(temporary)
            alternate = run_dir / "benchmark_results_old_kilitli_kopya.xlsx"
            alternate.write_bytes(b"old")
            pointer = {
                "latest_workbook": str(alternate),
                "requested_workbook": str(run_dir / "benchmark_results.xlsx"),
            }
            (run_dir / "excel_latest.json").write_text(json.dumps(pointer), encoding="utf-8")
            time.sleep(0.01)
            primary = run_dir / "benchmark_results.xlsx"
            primary.write_bytes(b"new")
            self.assertEqual(resolve_incremental_base(run_dir, primary), primary.resolve())

    def test_minimal_report_has_required_sheets_formulas_and_excel_safe_empty_table(self):
        with tempfile.TemporaryDirectory() as temporary:
            run_dir = Path(temporary)
            (run_dir / "run_config.json").write_text(
                json.dumps({"run_id": "unit_excel", "seed": 42}), encoding="utf-8"
            )
            summary = [
                {
                    "direction": "A__TO__B",
                    "query_variant": "hard_v1",
                    "search_mode": "global",
                    "model_id": "RAW_BASELINE",
                    "total_queries": 1,
                    "ok_queries": 1,
                    "rejected_queries": 0,
                    "error_queries": 0,
                    "coverage": 1.0,
                    "mean_error_m": 2.5,
                    "median_error_m": 2.5,
                    "success_25m_queries": 1,
                    "success_25m": 1.0,
                    "success_25m_failure_rate": 0.0,
                    "auc_25m": 0.9,
                    "mean_error_under_25m": 2.5,
                    "median_error_under_25m": 2.5,
                    "success_10m": 1.0,
                }
            ]
            (run_dir / "summary.json").write_text(json.dumps(summary), encoding="utf-8")
            result = {
                "run_id": "unit_excel",
                "status": "ok",
                "direction": "A__TO__B",
                "query_variant": "hard_v1",
                "search_mode": "global",
                "model_id": "RAW_BASELINE",
                "query_id": "Q00001",
                "error_m": 2.5,
            }
            (run_dir / "results.jsonl").write_text(json.dumps(result) + "\n", encoding="utf-8")

            output = run_dir / "benchmark_results.xlsx"
            report = build_workbook(run_dir, output)
            self.assertEqual(report["zip_integrity"], "ok")
            self.assertEqual(report["formula_error_literals"], [])

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                [name for name in workbook.sheetnames if name != "_ExcelState"],
                ["Özet", "Model Özeti", "Ham Sonuçlar", "Sorgu Manifesti", "Yapılandırma", "Hatalar"],
            )
            self.assertEqual(workbook["_ExcelState"].sheet_state, "veryHidden")
            self.assertEqual(workbook["Özet"]["B5"].value, "=COUNTA('Ham Sonuçlar'!A2:A2)")
            self.assertEqual(
                [workbook["Özet"].cell(14, column).value for column in range(1, 5)],
                ["Model (global arama)", "25 m başarı", "25 m içi medyan hata", "AUC@25m"],
            )
            self.assertIn("success_25m", SUMMARY_COLUMNS)
            self.assertIn("auc_25m", SUMMARY_COLUMNS)
            self.assertIn("median_error_under_25m", SUMMARY_COLUMNS)
            self.assertFalse(any("30m" in column for column in SUMMARY_COLUMNS))
            self.assertEqual(len(workbook["Özet"]._charts), 1)
            self.assertEqual(workbook["Hatalar"].max_row, 2)
            self.assertEqual(len(workbook["Hatalar"].tables), 1)

    def test_lightweight_checkpoint_omits_raw_data_and_ranks_only_global_search(self):
        with tempfile.TemporaryDirectory() as temporary:
            run_dir = Path(temporary)
            (run_dir / "run_config.json").write_text(
                json.dumps({"run_id": "lightweight_excel", "seed": 42}), encoding="utf-8"
            )
            base = {
                "direction": "A__TO__B",
                "query_variant": "clean",
                "model_id": "model_01",
                "total_queries": 10,
                "ok_queries": 9,
                "rejected_queries": 1,
                "error_queries": 0,
                "coverage": 0.9,
                "success_25m_queries": 8,
                "success_25m_failure_rate": 0.2,
                "median_error_under_25m": 4.0,
            }
            summary = [
                {**base, "search_mode": "global", "success_25m": 0.8, "auc_25m": 0.7},
                {**base, "search_mode": "roi_500m", "success_25m": 1.0, "auc_25m": 0.95},
            ]
            (run_dir / "summary.json").write_text(json.dumps(summary), encoding="utf-8")
            (run_dir / "results.jsonl").write_text(
                "".join(json.dumps({"status": "ok", "row": index}) + "\n" for index in range(500)),
                encoding="utf-8",
            )

            output = run_dir / "benchmark_results.xlsx"
            report = build_workbook(
                run_dir, output, validation_mode="checkpoint", lightweight=True
            )
            self.assertEqual(report["report_scope"], "lightweight")
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Özet", "Model Özeti", "Yapılandırma", "Hatalar"])
            self.assertNotIn("Ham Sonuçlar", workbook.sheetnames)
            self.assertEqual(
                workbook["Özet"]["B5"].value,
                "=SUM('Model Özeti'!E2:E3)",
            )
            self.assertIn("'Model Özeti'!D2", workbook["Özet"]["A15"].value)
            self.assertIsNone(workbook["Özet"]["A16"].value)
            self.assertIn("yalnız global aramayı", workbook["Özet"]["A10"].value)
            self.assertIn("ham sonuçlar yoktur", workbook["Özet"]["A10"].value)

    def test_incremental_update_appends_only_new_raw_rows_and_refreshes_dashboard(self):
        with tempfile.TemporaryDirectory() as temporary:
            run_dir = Path(temporary)
            (run_dir / "run_config.json").write_text(
                json.dumps({"run_id": "incremental_excel", "seed": 42}), encoding="utf-8"
            )
            first_summary = [
                {
                    "direction": "A__TO__B",
                    "query_variant": "clean",
                    "search_mode": "global",
                    "model_id": "RAW_BASELINE",
                    "total_queries": 1,
                    "ok_queries": 1,
                    "rejected_queries": 0,
                    "error_queries": 0,
                    "coverage": 1.0,
                    "success_25m_queries": 1,
                    "success_25m": 1.0,
                    "success_25m_failure_rate": 0.0,
                    "auc_25m": 0.9,
                    "median_error_under_25m": 2.5,
                }
            ]
            (run_dir / "summary.json").write_text(json.dumps(first_summary), encoding="utf-8")
            rows = [
                {
                    "run_id": "incremental_excel",
                    "status": "ok",
                    "direction": "A__TO__B",
                    "query_variant": "clean",
                    "search_mode": "global",
                    "model_id": "RAW_BASELINE",
                    "query_id": "Q00001",
                    "error_m": 2.5,
                }
            ]
            results_path = run_dir / "results.jsonl"
            results_path.write_text(
                "".join(json.dumps(row) + "\n" for row in rows), encoding="utf-8"
            )
            output = run_dir / "benchmark_results.xlsx"
            build_workbook(run_dir, output, validation_mode="checkpoint")

            rows.append(
                {
                    "run_id": "incremental_excel",
                    "status": "ok",
                    "direction": "A__TO__B",
                    "query_variant": "clean",
                    "search_mode": "global",
                    "model_id": "model_01",
                    "query_id": "Q00001",
                    "error_m": 4.0,
                }
            )
            results_path.write_text(
                "".join(json.dumps(row) + "\n" for row in rows), encoding="utf-8"
            )
            second_summary = first_summary + [
                {
                    **first_summary[0],
                    "model_id": "model_01",
                    "median_error_under_25m": 4.0,
                }
            ]
            (run_dir / "summary.json").write_text(json.dumps(second_summary), encoding="utf-8")

            report = update_workbook_incremental(
                run_dir, output, validation_mode="checkpoint"
            )
            self.assertTrue(report["incremental"])
            self.assertEqual(report["appended_result_rows"], 1)
            workbook = load_workbook(output, data_only=False)
            raw = workbook["Ham Sonuçlar"]
            self.assertEqual(raw.max_row, 3)
            self.assertEqual(raw.tables["RawResults_01"].ref.split(":")[-1][-1], "3")
            self.assertEqual(workbook["Özet"]["B5"].value, "=COUNTA('Ham Sonuçlar'!A2:A3)")
            self.assertEqual(len(workbook["Özet"]._charts), 1)
            self.assertEqual(workbook["Model Özeti"].max_row, 3)

            full_output = run_dir / "benchmark_results_full_reference.xlsx"
            build_workbook(run_dir, full_output, validation_mode="checkpoint")
            full_workbook = load_workbook(full_output, data_only=False)
            visible_names = [name for name in workbook.sheetnames if name != "_ExcelState"]
            self.assertEqual(
                visible_names,
                [name for name in full_workbook.sheetnames if name != "_ExcelState"],
            )
            for sheet_name in visible_names:
                incremental_sheet = workbook[sheet_name]
                full_sheet = full_workbook[sheet_name]
                self.assertEqual(
                    list(incremental_sheet.values),
                    list(full_sheet.values),
                    sheet_name,
                )
                self.assertEqual(
                    {
                        name: incremental_sheet.tables[name].ref
                        for name in incremental_sheet.tables
                    },
                    {name: full_sheet.tables[name].ref for name in full_sheet.tables},
                    sheet_name,
                )
                self.assertEqual(len(incremental_sheet._charts), len(full_sheet._charts), sheet_name)
            self.assertEqual(
                workbook["Ham Sonuçlar"]["H2"].number_format,
                full_workbook["Ham Sonuçlar"]["H2"].number_format,
            )
            self.assertEqual(
                workbook["Ham Sonuçlar"]["A1"].fill.fgColor.rgb,
                full_workbook["Ham Sonuçlar"]["A1"].fill.fgColor.rgb,
            )

    def test_locked_primary_is_preserved_and_a_timestamped_copy_is_created(self):
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            output = directory / "benchmark_results.xlsx"
            original_bytes = b"excel-is-open"
            output.write_bytes(original_bytes)
            workbook = Workbook()
            workbook.active["A1"] = "new report"
            real_replace = __import__("os").replace
            calls = 0

            def replace_with_primary_lock(source, destination):
                nonlocal calls
                calls += 1
                if Path(destination) == output and calls == 1:
                    raise PermissionError("simulated Excel lock")
                return real_replace(source, destination)

            with patch(
                "build_benchmark_excel_openpyxl.os.replace",
                side_effect=replace_with_primary_lock,
            ):
                actual, mode = save_workbook_safely(workbook, output)

            self.assertEqual(mode, "locked_copy")
            self.assertNotEqual(actual, output)
            self.assertIn("kilitli_kopya", actual.name)
            self.assertEqual(output.read_bytes(), original_bytes)
            self.assertTrue(actual.is_file())

    @unittest.skipUnless(sys.platform == "win32", "Windows file-sharing lock test")
    def test_real_windows_excel_style_lock_creates_copy_without_touching_primary(self):
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            output = directory / "benchmark_results.xlsx"
            original_workbook = Workbook()
            original_workbook.active["A1"] = "open in Excel"
            original_workbook.save(output)
            original_hash = output.read_bytes()

            create_file = ctypes.windll.kernel32.CreateFileW
            create_file.argtypes = [
                ctypes.c_wchar_p,
                ctypes.c_uint32,
                ctypes.c_uint32,
                ctypes.c_void_p,
                ctypes.c_uint32,
                ctypes.c_uint32,
                ctypes.c_void_p,
            ]
            create_file.restype = ctypes.c_void_p
            handle = create_file(str(output), 0x80000000, 0x00000001, None, 3, 0x80, None)
            self.assertNotEqual(handle, ctypes.c_void_p(-1).value)
            try:
                updated_workbook = Workbook()
                updated_workbook.active["A1"] = "new checkpoint"
                actual, mode = save_workbook_safely(updated_workbook, output)
            finally:
                ctypes.windll.kernel32.CloseHandle(ctypes.c_void_p(handle))

            self.assertEqual(mode, "locked_copy")
            self.assertNotEqual(actual, output)
            self.assertEqual(output.read_bytes(), original_hash)
            self.assertEqual(load_workbook(actual).active["A1"].value, "new checkpoint")


if __name__ == "__main__":
    unittest.main()
